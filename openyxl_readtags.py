from openpyxl import load_workbook
# 使用pandas读取excel文件
import unittest
from adbtest.utils import util
from adbtest.devices_key import samsung6_inputkey
import time
from appium import webdriver

desired_caps = {};
desired_caps['platformName'] = 'Android'
desired_caps['platformVersion'] = '6.0.1'
desired_caps['deviceName'] = '5203adddfc7334c1'
desired_caps['noReset'] = 'true'
oneplus_desired_caps = {};
oneplus_desired_caps['platformName'] = 'Android'
oneplus_desired_caps['platformVersion'] = '10'
oneplus_desired_caps['deviceName'] = '4e62be76'
oneplus_desired_caps['noReset'] = 'true'

# driver = webdriver.Remote('http://localhost:4723/wd/hub', desired_caps);


class en_correctTest(unittest.TestCase):
    def setUp(self):
        self.driver = webdriver.Remote('http://localhost:4723/wd/hub', desired_caps);

    def tearDown(self):
        self.driver.quit()

    def test_start(self):
        file_path = r'\Users\xm20190901\Downloads\tag.xlsx'
        data = load_workbook(r'/Users/xm20190901/Downloads/tag.xlsx')
        ws = data.active  # 获取工作的激活工作表
        sheet = data['Sheet1']
        rows = sheet.max_row - 1
        row_info = sheet.cell(row=2, column=1).value  # 读取第二行第一列内容
        for irow in range(rows):
            print(irow)
            tag = sheet.cell(row=irow + 2, column=1).value  # 从第二行第一列开始读取tag
            print(tag)
            for i in range(len(tag) + 1):
                str_tag = tag[i - 1:i]  # 读取tag每一个字符
                print(str_tag)
                time.sleep(1)
                utils.input_by_tag(str_tag, self)  # 根据tag每一个字符按下对应键盘按键
            samsung6_inputkey.space(self, 0)
            message = utils.gettext(self)
            if message.lower() == tag: #将输入框内容message转换成小写字母和tag对比
                sheet.cell(row=irow + 2, column=3).value = "输入内容和tag一致"
                time.sleep(1)
                data.save(r'/Users/xm20190901/Downloads/tag.xlsx')
                print("保存通过")
            else:
                sheet.cell(row=irow + 2, column=3).value = "writefail"
                time.sleep(1)
                data.save(r'/Users/xm20190901/Downloads/tag.xlsx')
                print("保存失败")
            time.sleep(1)
            utils.cleartext(self)


if __name__ == '__main__':
    # 构造测试集  defaultTestLoader（）即TestLoader（）测试用例加载器，包括多个加载测试用例的方法，返回一个测试套件
    # loadTestsFromTestCase（）根据给定的测试类，获取其中的所有测试方法，并返回一个测试套件
    suite = unittest.TestLoader().loadTestsFromTestCase(en_correctTest)
    # unittest框架的TextTestRunner（）类，通过该类下面的run（）方法来运行suite所组装的测试用例，入参为suite测试套件
    unittest.TextTestRunner(verbosity=2).run(suite)
    # 上面两行代码可以换成下面一行
    # unittest.main()
